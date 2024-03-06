import open3d as o3d
import numpy as np
from sklearn import decomposition as decomp
from matplotlib import pyplot as plt
from matplotlib import cm
import random
import copy
import os
import glob
import pandas as pd
from sklearn.neighbors import NearestNeighbors
from openpyxl import Workbook, load_workbook
import openpyxl.styles as styles
from datetime import datetime

class PointCloudStructure: 
    def __init__(self, pointCloud: o3d.geometry.PointCloud = None, pointCloudFilePath: str = None, stl_path: str = None, stl_sample_points: int = None, show = False) -> None:
        """Creates a point cloud structure. This structure can be created from either a point cloud, a point cloud file path, or an STL file path. 
        If you provide an STL file path, you must also provide the number of sample points to use for the STL file. 
        

        Args:
            pointCloud (o3d.geometry.PointCloud, optional): Open3d point cloud to use in the creation of this datatype. Defaults to None.
            pointCloudFilePath (str, optional): Relative filepath to the point cloud to use to create thise datatype. Defaults to None.
            stl_path (str, optional): Relative filepath to the STL file to use to create this datatype. Defaults to None.
            stl_sample_points (int, optional): Number of sample points to use for the STL file. Defaults to None.
            show (bool, optional): Whether or not to show the original point cloud. Defaults to False.

        Raises:
            Exception: If you provide more than one of a point cloud, a file path, or an STL, it will raise an exception.
            Exception: If you provide an STL file path without a number of sample points, it will raise an exception.
        """
        
        if (pointCloud != None and pointCloudFilePath != None) or (stl_path != None and pointCloud != None) or (stl_path != None and pointCloudFilePath != None) or (stl_path==None and pointCloud==None and pointCloudFilePath==None):
            raise Exception("Please provide either a pointCloud, a pointCloudFilePath, or an STL filepath.")
        
        if stl_path != None and stl_sample_points == None:
            raise Exception("Please provide a number of sample points for the STL file.")
        
        if stl_path != None:
            self.point_cloud_filepath = os.path.join(os.getcwd(), stl_path)
            mesh = o3d.io.read_triangle_mesh(stl_path)
            self.PointCloud = mesh.sample_points_poisson_disk(stl_sample_points)
        elif pointCloud != None:
            self.point_cloud_filepath = None
            self.PointCloud = pointCloud
        elif pointCloudFilePath != None:
            self.point_cloud_filepath = os.path.join(os.getcwd(), pointCloudFilePath)
            self.PointCloud = o3d.io.read_point_cloud(pointCloudFilePath)
        self.noisy_cloud = o3d.geometry.PointCloud()
        self.cleaned_cloud = o3d.geometry.PointCloud()
        self.original_cloud_average_distance = np.mean(o3d.geometry.PointCloud.compute_nearest_neighbor_distance(self.PointCloud))
        self.original_number_of_points = len(self.PointCloud.points)
        original_PCA = decomp.PCA(n_components=3)
        original_PCA.fit(np.asarray(self.PointCloud.points))
        self.original_lowest_variance_ratio = original_PCA.explained_variance_ratio_[2]
        self.viewing_angle = None
        
        if show:
            print("Showing original point cloud")
            self.PointCloud.paint_uniform_color([0.4, 0.4, 0.6])
            o3d.visualization.draw_geometries([self.PointCloud])
    
    def generate_statistical_removal_and_PCA_variance_data(self, nb_neighbor_upper: int, nb_neighbors_samples: int, std_dev_upper: float,
                                               std_dev_samples: int, nb_neighbor_lower: int = 1,
                                               std_dev_lower: float = .000001, graph: bool = False) -> None:
        """Generates data regarding the variance of the principal components as a function of statistical outlier removal. Stores this data and the parameters that give the least variance ratio.
        This data is stored to be used later in the create_summary_subplots function. 

        Args:
            nb_neighbor_upper (int): The upper bound for the number of neighbors argument.
            nb_neighbors_samples (int): The number of samples to test for nb_neighbors.
            std_dev_upper (float): The upper bound for the standard deviation argument.
            std_dev_samples (int): The number of samples to use for standard deviation.
            nb_neighbor_lower (int, optional): The lower bound for the number of neighbors argument. Defaults to 1.
            std_dev_lower (float, optional): The lower bound for the standard deviation argument. Defaults to .000001.
            graph (bool, optional): Whether or not to graph the data. Defaults to False.

        Raises:
            Exception: If you try to run this function without generating a noisy cloud first, it will raise an exception.
        """
        
        
        
        if self.noisy_cloud == None:
            raise Exception("No noisy cloud has been generated. Please generate a noisy cloud before calling\
                            this function, by running PointCloud.generate_gaussian_noise.")
        
        self.best_least_variance_ratio = np.inf
        
        pcd = copy.copy(self.noisy_cloud)
        self.nb_neighbor_points_pca = np.linspace(nb_neighbor_lower, nb_neighbor_upper, nb_neighbors_samples).astype(int)
        self.std_dev_points_pca = np.linspace(std_dev_lower, std_dev_upper, std_dev_samples)
        self.explained_variance_ratio_data = np.empty((3, self.nb_neighbor_points_pca.size, self.std_dev_points_pca.size))
        for i in range(self.nb_neighbor_points_pca.size):
            for j in range(self.std_dev_points_pca.size):
                downpcd, _ = pcd.remove_statistical_outlier(nb_neighbors=self.nb_neighbor_points_pca[i],
                                                            std_ratio=self.std_dev_points_pca[j])
                points = np.asarray(downpcd.points)
                if points.size != 0:
                    PCA = decomp.PCA(n_components=3)
                    PCA.fit(points)
                    for k in range(3):
                        self.explained_variance_ratio_data[k, i, j] = PCA.explained_variance_ratio_[k]
                        if PCA.explained_variance_ratio_[k] < self.best_least_variance_ratio:
                            self.best_least_variance_ratio = PCA.explained_variance_ratio_[k]
                            self.best_nb_neighbors_PCA = self.nb_neighbor_points_pca[i]
                            self.best_std_dev_PCA = self.std_dev_points_pca[j]
                            self.best_PCA_cleaned_cloud = downpcd
                            self.best_PCA_filter_number_of_points = len(downpcd.points)
                else:
                    for k in range(3):
                        self.explained_variance_ratio_data[k, i, j] = 'nan'
        self.nb_neighbor_points_pca, self.std_dev_points_pca = np.meshgrid(self.nb_neighbor_points_pca, self.std_dev_points_pca)
        graph_labels_ratio = ['Principal Component Highest Variance Ratio', 'Principal Component Medium Variance Ratio',
                            'Principal Component Least Variance Ratio']
        graph_labels_amount = ['Variance Highest', 'Variance Medium', 'Variance Lowest']
        
        
        
        if graph == True:
            ax = plt.axes(projection='3d', computed_zorder=False)
            ax.plot_surface(self.std_dev_points_pca, self.nb_neighbor_points_pca, np.transpose(self.explained_variance_ratio_data[2, :, :] * 10**3),
                                cmap=cm.coolwarm,
                                linewidth=0, antialiased=False)
            ax.scatter(self.best_std_dev_PCA, self.best_nb_neighbors_PCA, self.best_least_variance_ratio * 10**3, s=250, c='r', marker = '*', zorder = 100)
            ax.set_title(graph_labels_ratio[2])
            ax.set_xlabel('Standard ratio')
            ax.set_ylabel('Number of neighbors')
            ax.set_zlabel('Variance ratio * 1e3')
            ax.view_init(azim=160, elev=30)
            plt.show()
        return

    def generate_statistical_removal_and_average_distance_data(self, nb_neighbor_upper: int, nb_neighbors_samples: int, std_dev_upper: float,
                                                   std_dev_samples: int, nb_neighbor_lower: int = 1,
                                                   std_dev_lower: float = .000001, graph: bool = False) -> None:
        """Generates data regarding the average distance between points and their nearest neighbors
        as a function of statistical outlier removal. Stores this data and the parameters that give the least distance between points. 
        This data is stored to be used later in the create_summary_subplots function.

        Args:
            nb_neighbor_upper (int): Upper bound for number of neighbors argument
            nb_neighbors_samples (int): Number of samples to test for nb_neighbors
            std_dev_upper (float): Upper bound for standard deviation argument
            std_dev_samples (int): Number of samples to use for standard deviation
            nb_neighbor_lower (int, optional): Lower bound for number of neighbors argument. Default 1.
            std_dev_lower (float, optional): Lower bound for standard deviation argument. Default .000001
            graph (bool, optional): Whether or not to graph the data. Default False.

        Raises:
            Exception: If you try to run this function without generating a noisy cloud first, it will raise an exception.  
        """
        
        if self.noisy_cloud == None:
            raise Exception("No noisy cloud has been generated. Please generate a noisy cloud before calling\
                        this function, by running PointCloud.generate_gaussian_noise.")
        
        self.best_mean_distance = np.inf
        
        self.nb_neighbor_points_ad = np.linspace(nb_neighbor_lower, nb_neighbor_upper, nb_neighbors_samples).astype(int)
        self.std_dev_points_ad = np.linspace(std_dev_lower, std_dev_upper, std_dev_samples)
        self.average_distance_data = np.empty((self.nb_neighbor_points_ad.size, self.std_dev_points_ad.size))
        for i in range(self.nb_neighbor_points_ad.size):
            for j in range(self.std_dev_points_ad.size):
                downpcd, _ = self.noisy_cloud.remove_statistical_outlier(nb_neighbors=self.nb_neighbor_points_ad[i], std_ratio=self.std_dev_points_ad[j])
                points = np.asarray(downpcd.points)
                if points.size != 0:
                    mean_distance = np.mean(o3d.geometry.PointCloud.compute_nearest_neighbor_distance(downpcd))
                    if mean_distance < self.best_mean_distance:
                        self.best_mean_distance = mean_distance
                        self.best_nb_neighbors_mean = self.nb_neighbor_points_ad[i]
                        self.best_std_dev_mean = self.std_dev_points_ad[j]
                        self.best_average_distance_cleaned_cloud = downpcd
                        self.best_average_distance_filter_number_of_points = len(downpcd.points)
                    self.average_distance_data[i, j] = mean_distance
                else:
                    self.average_distance_data[i, j] = 'nan'
        self.nb_neighbor_points_ad, self.std_dev_points_ad = np.meshgrid(self.nb_neighbor_points_ad, self.std_dev_points_ad)
        
        if graph == True:
            
            ax = plt.axes(projection='3d')
            ax.plot_surface(self.nb_neighbor_points_ad, self.std_dev_points_ad, np.transpose(self.average_distance_data) * 10**4,
                                cmap=cm.coolwarm,
                                linewidth=0, antialiased=False)
            ax.set_title('Nearest Neighbor Average Distance')
            ax.set_xlabel('Number of neighbors')
            ax.set_ylabel('Standard ratio')
            ax.set_zlabel('Average distance * 1e4')
            ax.scatter(self.best_nb_neighbors_mean, self.best_std_dev_mean, self.best_mean_distance * 10**4, s=250, c='r', marker = '*', zorder = 100)
            ax.yaxis.labelpad=10
            ax.xaxis.labelpad=10
            ax.zaxis.labelpad=10
            ax.xaxis.set_major_locator(plt.MaxNLocator(5))
            ax.zaxis.set_major_locator(plt.MaxNLocator(5))
            plt.show()
        return
    
    def display_noise_not_noise(self, cloud, ind):
            noise_cloud = cloud.select_by_index(ind)
            correct_cloud = cloud.select_by_index(ind, invert=True)
            noise_cloud.paint_uniform_color([1, 0, 0])
            correct_cloud.paint_uniform_color([0.4, 0.4 , 0.6])
            print("Showing noissy point cloud.")
            print('Noise points are shown in red, correct points are shown in gray.')
            # if type(self.viewing_angle) != type(None):
            #     o3d.visualization.draw_geometries([noise_cloud, correct_cloud], view_control=o3d.visualization.ViewControl.convert_from_pinhole_camera_parameters(self.viewing_angle))
            o3d.visualization.draw_geometries([noise_cloud, correct_cloud])
            #self.viewing_angle = o3d.visualization.ViewControl.convert_to_pinhole_camera_parameters()
    
    def generate_gaussian_noise(self, mean: float, bounding_box_mult: float, noise_percent: float, show: bool = False) -> None:
        """Generates Gaussian distributed noise for a point cloud.

        Args:
            mean (float):  Mean distance the noise moves from the original point.
            bounding_box_mult (float): Standard deviation for the normal distribution that decides how much to move
            a true point by. Measured by taking this argument and multiplying it by the square root of the sum
            of the squares of the bounding box sizes of the original point cloud
            noise_percent (float): What percentage of points are moved from their original location.
            show (bool, optional): Whether or not to show the noisy point cloud, with noise points in red and correct points in gray. Defaults to False.
        """    
        
        self.noise_indeces = []
        points = np.array([1, 1])
        noise_counter = 0
        pcd = self.PointCloud
        points = copy.copy(np.asarray(self.PointCloud.points))
        original_max_bounds = self.PointCloud.get_max_bound()
        original_min_bounds = self.PointCloud.get_min_bound()
        self.original_ranges = np.asarray(original_max_bounds) - np.asarray(original_min_bounds)
        bounding_box_magnitude = np.sqrt(np.sum(self.original_ranges ** 2))
        std_dev = bounding_box_magnitude*bounding_box_mult
        for i in range(len(points)):
            if random.random() <= noise_percent:
                self.noise_indeces.append(i)
                noise_counter += 1
                noise = np.random.normal(mean, std_dev, (1, 3))
                points[i] = points[i] + noise
        self.noisy_cloud = o3d.geometry.PointCloud()
        self.noisy_cloud.points = o3d.utility.Vector3dVector(points)
        self.noise_counter = noise_counter
        if show:
            self.display_noise_not_noise(self.noisy_cloud, self.noise_indeces)
        
    def chamfer_distance(self, x: o3d.geometry.PointCloud, y: o3d.geometry.PointCloud) -> float:
        """Calculates the chamfer distance between two point clouds, using a k-nearest-neighbor data structure.

        Args:
            x (o3d.geometry.PointCloud): Point cloud 1
            y (o3d.geometry.PointCloud): Point cloud 2
        Returns:
            float: Chamfer distance between the two point clouds.
        """
    
        x = np.asarray(x.points)
        y = np.asarray(y.points)
    
        x_nn = NearestNeighbors(n_neighbors=1, leaf_size=1, algorithm='kd_tree').fit(x)
        min_y_to_x = x_nn.kneighbors(y)[0]
        y_nn = NearestNeighbors(n_neighbors=1, leaf_size=1, algorithm='kd_tree').fit(y)
        min_x_to_y = y_nn.kneighbors(x)[0]
        chamfer_dist = np.mean(min_y_to_x) + np.mean(min_x_to_y)
            
        return chamfer_dist
        
    def evaluate_filter_parameters(self, evaluation_method: str, nb_neighbors: int = 0, std_ratio: float = 0, method:str = None, show_images:bool = False, reduce_to_summary_statistic: bool = True) -> (float):
        """Evaluate the filter quality for the statistical outlier removal filter. If the evaluation method is "Summary statistic", and you set reduce_to_summary_statistic
        to false, reports accuracy, recall, precision, and Normalized RMSE = RMSE/(average distance between points in the original cloud). If reduce_to_summary_statistic
        is set to True, reports a summary statistic which is average(precision, recall, accuracy)*(average distance between points in the original cloud)/(RMSE).
        If you provide a method instead of nb_neighbord and std_ratio, it will automatically select the hypothetically "best" nb_neighbors and std_ratio based on this 
        method, by finding the parameters that create a minimum for the method specified. Finally, if evaluation method is set to "Chamfer distance", it will 
        return the chamfer distance between the original point cloud and the filtered point cloud.

        Args:
            evaluation_method (str): Either "Summary statistic" or "Chamfer distance". If you select "Summary statistic", you can also select whether or not to reduce the output to a single summary statistic.
            nb_neighbors (int, optional): nb_neighbors argument to use for the statistical outlier removal filter. Defaults to 0.
            std_ratio (int, optional): std_ratio argument to use for the statistical outlier removal filter. Defaults to 0.
            method (str, optional): Use either 'PCA' or 'Average distance'. This will automatically select the hypothetically 'best'
            filter parameters using the selected method. Defaults to None.
            show_images (bool, optional): Whether or not to show the original, noisy, and filtered point clouds. Defaults to False.
            reduce_to_summary_statistic (bool, optional): Whether or not to reduce the output to a single summary statistic. Defaults to True.
            
        Returns:
            float: Summary statistic for the filter quality. This is defined as average(precision, recall, accuracy)*(average distance between points in the original cloud)/(RMSE)
            or
            float: Chamfer distance between the original point cloud and the filtered point cloud.
            or
            tuple: (recall, accuracy, noise detection rate, RMSE) if reduce_to_summary_statistic is set to False.
            
        Raises:
            Exception: If you run this function without providing either nb_neighbors and std_ratio or method, it will raise an exception.
            Exception: If you provide an evaluation method other than 'Summary statistic' or 'Chamfer distance', it will raise an exception.
        """ 
               
        if evaluation_method != "Chamfer distance" and evaluation_method != "Summary statistic":
            raise Exception("Please provide either 'Chamfer distance' or 'Summary statistic' for the evaluation method.")
               
        if (nb_neighbors == 0 or std_ratio == 0) and method == None:
            raise Exception('Please provide either nb_neighbors and std_ratio or method for filter evaluation.')
        
        if method == 'PCA':
            nb_neighbors = self.best_nb_neighbors_PCA
            std_ratio = self.best_std_dev_PCA
            
        elif method == 'Average distance':
            nb_neighbors = self.best_nb_neighbors_mean
            std_ratio = self.best_std_dev_mean
        
        pcd_in = copy.copy(self.PointCloud)
        pcd_noisy = copy.copy(self.noisy_cloud)
        clean_array = np.asarray(pcd_in.points)
        noisy_array = np.asarray(pcd_noisy.points)
        cleaned_cloud, ind = pcd_noisy.remove_statistical_outlier(nb_neighbors, std_ratio)
        if show_images == True:
            print("Showing original point cloud")
            o3d.visualization.draw_geometries([pcd_in])
            print("Showing noisy point cloud")
            o3d.visualization.draw_geometries([pcd_noisy])
            print("Showing filtered point cloud")
            o3d.visualization.draw_geometries([cleaned_cloud])
        
        if evaluation_method == "Summary statistic":
            noise_counter_filtered = 0
            for index in ind:
                if not (noisy_array[index] == clean_array[index]).all():
                        noise_counter_filtered += 1
            noise_points_removed = self.noise_counter -  noise_counter_filtered
            total_points_removed = self.original_number_of_points - len(ind)
            average_point_distance = np.mean(o3d.geometry.PointCloud.compute_nearest_neighbor_distance(self.PointCloud))
            if self.noise_counter == 0 or (noise_points_removed + total_points_removed == 0) or self.original_number_of_points == 0 or np.size(clean_array, 0) == 0 or average_point_distance == 0:
                return np.nan
            noise_detection_rate = noise_points_removed / self.noise_counter * 100
            correct_points_removed = total_points_removed - noise_points_removed
            correct_points_left = self.original_number_of_points - self.noise_counter - (total_points_removed - noise_points_removed)
            recall = noise_points_removed / (noise_points_removed + correct_points_removed) * 100
            accuracy = (noise_points_removed + correct_points_left) / (self.original_number_of_points) * 100
            RMSE = np.sqrt(np.sum((clean_array[ind] - cleaned_cloud.points) ** 2) / np.size(clean_array, 0))/average_point_distance
            if show_images == True:
                print('recall: ' + str(recall) + '%')
                print('accuracy: ' + str(accuracy) + '%')
                print('noise detection rate: ' + str(noise_detection_rate) + '%')
                print('Normalized RMSE: ' + str(RMSE))
            
            if RMSE != 0 and reduce_to_summary_statistic == True:
                if show_images == True:
                    print('Summary statistic: ' + str(np.mean([recall, np.float64(accuracy), noise_detection_rate]) / (RMSE)))
                return np.mean([recall, np.float64(accuracy), noise_detection_rate]) / (RMSE)
            
            elif reduce_to_summary_statistic == False and RMSE != 0:
                return recall, accuracy, noise_detection_rate, RMSE
            
            else:
                # print('RMSE was zero for nb_neighbors = ' + str(nb_neighbors) + ' and std_ratio = ' + str(std_ratio) + '.')
                return np.nan
            
        elif evaluation_method == "Chamfer distance":
            if len(cleaned_cloud.points) > 0:
                chamfer_distance = self.chamfer_distance(self.PointCloud, cleaned_cloud)
            else:
                chamfer_distance = np.nan
                
            return chamfer_distance
        
    def generate_summary_statistic_data(self, nb_neighbor_upper: int, nb_neighbors_samples: int, std_dev_upper: float,
                                               std_dev_samples: int, nb_neighbor_lower: int = 1,
                                               std_dev_lower: float = .000001, graph: bool = False) -> None:
        """Generate data for the summary statistic over the input domain. The summary statistic is defined as
        average(precision, recall, accuracy)*(average distance between points in the original cloud)/(RMSE).

        Args:
            nb_neighbor_upper (int): Upper bound for number of neighbors argument
            nb_neighbors_samples (int): Number of samples to test for nb_neighbors
            std_dev_upper (float): Upper bound for standard deviation argument
            std_dev_samples (int): Number of samples to use for standard deviation
            nb_neighbor_lower (int, optional): Lower bound for number of neighbors argument. Default 1.
            std_dev_lower (float, optional): Lower bound for standard deviation argument. Default .000001
            graph (bool, optional): Whether or not to graph the data. Default False.
            
        Raises:
            Exception: If you try to run this function without generating a noisy cloud first, it will raise an exception.
        """        
        if self.noisy_cloud == None:
            raise Exception("No noisy cloud has been generated. Please generate a noisy cloud before calling\
                        this function, by running PointCloud.generate_gaussian_noise.")
        
        self.best_summary_statistic = 0
        
        self.nb_neighbor_points_ss = np.linspace(nb_neighbor_lower, nb_neighbor_upper, nb_neighbors_samples).astype(int)
        self.std_dev_points_ss = np.linspace(std_dev_lower, std_dev_upper, std_dev_samples)
        self.summary_statistic_data = np.empty((self.nb_neighbor_points_ss.size, self.std_dev_points_ss.size))
        for i in range(self.nb_neighbor_points_ss.size):
            for j in range(self.std_dev_points_ss.size):
                self.summary_statistic_data[i, j] = self.evaluate_filter_parameters(evaluation_method="Summary statistic", nb_neighbors=self.nb_neighbor_points_ss[i], std_ratio=self.std_dev_points_ss[j])
                if self.summary_statistic_data[i, j] > self.best_summary_statistic:
                    self.best_summary_statistic = self.summary_statistic_data[i, j]
                    self.best_nb_neighbors_ss = self.nb_neighbor_points_ss[i]
                    self.best_std_dev_ss = self.std_dev_points_ss[j]
        self.nb_neighbor_points_ss, self.std_dev_points_ss = np.meshgrid(self.nb_neighbor_points_ss, self.std_dev_points_ss)
        self.optimal_filtered_cloud_ss, _ = self.noisy_cloud.remove_statistical_outlier(self.best_nb_neighbors_ss, self.best_std_dev_ss)
        if graph == True:
            print('Graphing summary statistic over input domain')
            ax = plt.axes(projection='3d')
            ax.plot_surface(self.nb_neighbor_points_ss, self.std_dev_points_ss, np.transpose(self.summary_statistic_data),
                                cmap=cm.coolwarm,
                                linewidth=0, antialiased=False)
            ax.set_title('Summary statistic over input domain')
            ax.set_xlabel('Number of neighbors')
            ax.set_ylabel('Standard deviation')
            ax.set_zlabel('Summary statistic')
            plt.show()
        return
    
    def generate_chamfer_distance_data(self, nb_neighbor_upper: int, nb_neighbors_samples: int, std_dev_upper: float,
                                               std_dev_samples: int, nb_neighbor_lower: int = 1,
                                               std_dev_lower: float = .000001, graph: bool = False) -> None:
        """Generate data for the chamfer distance over the input domain. Measures chamfer distance from filtered cloud to ground truth cloud.

        Args:
            nb_neighbor_upper (int): Upper bound for number of neighbors argument
            nb_neighbors_samples (int): Number of samples to test for nb_neighbors
            std_dev_upper (float): Upper bound for standard deviation argument
            std_dev_samples (int): Number of samples to use for standard deviation
            nb_neighbor_lower (int, optional): Lower bound for number of neighbors argument. Default 1.
            std_dev_lower (float, optional): Lower bound for standard deviation argument. Default .000001
            graph (bool, optional): Whether or not to graph the data. Default False.
            
        Raises:
            Exception: If you try to run this function without generating a noisy cloud first, it will raise an exception.
        """        
        if self.noisy_cloud == None:
            raise Exception("No noisy cloud has been generated. Please generate a noisy cloud before calling\
                        this function, by running PointCloud.generate_gaussian_noise.")
        
        self.best_chamfer_distance = np.inf
        
        self.nb_neighbor_points_pca_chamfer_distance = np.linspace(nb_neighbor_lower, nb_neighbor_upper, nb_neighbors_samples).astype(int)
        self.std_dev_points_pca_chamfer_distance = np.linspace(std_dev_lower, std_dev_upper, std_dev_samples)
        self.chamfer_distance_data = np.empty((self.nb_neighbor_points_pca_chamfer_distance.size, self.std_dev_points_pca_chamfer_distance.size))
        for i in range(self.nb_neighbor_points_pca_chamfer_distance.size):
            for j in range(self.std_dev_points_pca_chamfer_distance.size):
                self.chamfer_distance_data[i, j] = self.evaluate_filter_parameters(evaluation_method="Chamfer distance", nb_neighbors=self.nb_neighbor_points_pca_chamfer_distance[i], std_ratio=self.std_dev_points_pca_chamfer_distance[j])
                if self.chamfer_distance_data[i, j] < self.best_chamfer_distance:
                    self.best_chamfer_distance = self.chamfer_distance_data[i, j]
                    self.best_nb_neighbors_chamfer_distance = self.nb_neighbor_points_pca_chamfer_distance[i]
                    self.best_std_dev_chamfer_distance = self.std_dev_points_pca_chamfer_distance[j]
        self.nb_neighbor_points_pca_chamfer_distance, self.std_dev_points_pca_chamfer_distance = np.meshgrid(self.nb_neighbor_points_pca_chamfer_distance, self.std_dev_points_pca_chamfer_distance)
        
        self.optimal_filtered_cloud_chamfer_distance, _ = self.noisy_cloud.remove_statistical_outlier(self.best_nb_neighbors_chamfer_distance, self.best_std_dev_chamfer_distance)
        
        if graph == True:
            print('Graphing summary statistic over input domain')
            ax = plt.axes(projection='3d')
            ax.plot_surface(self.nb_neighbor_points_pca_chamfer_distance, self.std_dev_points_pca_chamfer_distance, np.transpose(self.chamfer_distance_data),
                                cmap=cm.coolwarm,
                                linewidth=0, antialiased=False)
            ax.set_title('Chamfer distance over input domain')
            ax.set_xlabel('Number of neighbors')
            ax.set_ylabel('Standard deviation')
            ax.set_zlabel('Chamfer distance')
            plt.show()
        return

    def create_summary_subplots(self, method: str, show: bool = False, save_path: str = False, name: str = None, number: int = 1):
        """Creates a matplotlib figure with 8 subplots. The first subplot is the original point cloud. The second subplot is the noisy point cloud.
        The third subplot is a graph representing output from either the average distance method or 
        the PCA method, depending on which method is selected. The fourth subplot is a graph
        representing the summary statistic. The fifth subplot is the filtered point cloud using parameters that provide a minimum for the
        method selected. The sixth subplot is the filtered point cloud using the optimal parameters for the summary statistic. The seventh subplot 
        is the filtered point cloud using the optimal parameters for the chamfer distance. The eighth subplot is the chamfer distance over the input domain.

        Args:
            method (str): Method for graphing the third subplot. Either 'PCA' or 'Average distance'.
            show (bool, optional): Whether or not to show the figure. Defaults to False.
            save_path (str, optional): If you want to save the figure, provide a filepath. Defaults to False.
            name (str, optional): If you want to save the figure, provide a name. Defaults to None.

        Raises:
            Exception: If you try to run this function without selecting a method, it will raise an exception.
            Exception: If you try to save the figure without providing a name, it will raise an exception.
        """        
        if method != 'PCA' and method != 'Average distance':
            raise Exception('method must be either "PCA" or "Average distance"')
        
        if save_path != False and type(name) == type(None):
            raise Exception ('If you are saving this point cloud\'s summary subplots, please give it a name.')
        
        fig = plt.figure(figsize=(10, 7))
        ax = fig.add_subplot(2,4,1, projection='3d')
        plt.rcParams.update({'font.size': 9})
        
        labelpad = 10
        
        
        #Graphing original point cloud
        
        (ax).set_title('Original Point Cloud')
        point_cloud_points = np.asarray(self.PointCloud.points)
        (ax).scatter(point_cloud_points[:,0], point_cloud_points[:,1], point_cloud_points[:,2], s=.5)
        ax.set_aspect('equal', adjustable='box')
        ax.set_yticklabels([])
        ax.set_xticklabels([])
        ax.set_zticklabels([])
        ax.view_init(azim=70, elev=20)
        
        #Graphing noisy point cloud
        
        ax = fig.add_subplot(2,4,2, projection='3d')
        (ax).set_title('Noisy Point Cloud')
        point_cloud_points = np.asarray(self.noisy_cloud.points)
        (ax).scatter(point_cloud_points[self.noise_indeces,0], point_cloud_points[self.noise_indeces,1], point_cloud_points[self.noise_indeces,2], s=.01, c='r')
        noise_mask = np.ones(len(point_cloud_points), dtype=bool)
        noise_mask[self.noise_indeces] = False
        non_noise = point_cloud_points[noise_mask]
        (ax).scatter(non_noise[: ,0], non_noise[: ,1], non_noise[: ,2], s=.5)
        ax.set_aspect('equal', adjustable='box')
        ax.set_yticklabels([])
        ax.set_xticklabels([])
        ax.set_zticklabels([])
        ax.view_init(azim=70, elev=20)
        
        #Graphing minimum of PCA method of filtering
        
        if method == 'PCA':
        
            ax = fig.add_subplot(2,4,3, projection='3d', computed_zorder=False)
            ax.set_title('Principal Component\nLeast Variance Ratio')
            ax.plot_surface(self.nb_neighbor_points_pca, self.std_dev_points_pca, np.transpose(self.explained_variance_ratio_data[2, :, :] * 10**3),
                                    cmap=cm.coolwarm,
                                    linewidth=0, antialiased=False, zorder = 0)
            #ax.scatter(self.best_nb_neighbors_PCA, self.best_std_dev_PCA, self.best_least_variance_ratio * 10**3, s=50, c='r', marker = '*', zorder = 100)
            ax.set_ylabel('Standard ratio')
            ax.set_xlabel('Number of neighbors')
            ax.set_zlabel('Variance ratio * 1e3')
            ax.view_init(azim=290, elev=30)
            ax.yaxis.labelpad=labelpad
            ax.xaxis.labelpad=labelpad
            ax.zaxis.labelpad=labelpad
            ax.yaxis.set_major_locator(plt.MaxNLocator(5))
            ax.zaxis.set_major_locator(plt.MaxNLocator(5))
            
        #Graphing minimum of average distance method of filtering    
            
        elif method == 'Average distance':
        
            ax = fig.add_subplot(2,4,3, projection='3d', computed_zorder=False)
            ax.plot_surface(self.nb_neighbor_points_ad, self.std_dev_points_ad, np.transpose(self.average_distance_data) * 10**4,
                                cmap=cm.coolwarm,
                                linewidth=0, antialiased=False, zorder = 0)
            ax.scatter(self.best_nb_neighbors_mean, self.best_std_dev_mean, self.best_mean_distance * 10**4, s=50, c='r', marker = '*', zorder = 100)
            ax.set_title('Nearest Neighbor Average Distance')
            ax.set_xlabel('Number of neighbors')
            ax.set_ylabel('Standard ratio')
            ax.set_zlabel('Average distance * 1e4')
            ax.view_init(azim=290, elev=30)
            ax.yaxis.labelpad=labelpad
            ax.xaxis.labelpad=labelpad
            ax.zaxis.labelpad=labelpad
            ax.xaxis.set_major_locator(plt.MaxNLocator(5))
            ax.zaxis.set_major_locator(plt.MaxNLocator(5))
            
        #Graphing summary statistic over input domain
            
        ax = fig.add_subplot(2,4,4, projection='3d', computed_zorder=False)
        ax.set_title('Summary Statistic\n over Input Domain')
        ax.plot_surface(self.nb_neighbor_points_ss, self.std_dev_points_ss, np.transpose(self.summary_statistic_data),
                            cmap=cm.coolwarm,
                            linewidth=0, antialiased=False)
        ax.set_title('Summary statistic over input domain')
        ax.set_xlabel('Number of neighbors')
        ax.set_ylabel('Standard ratio')
        ax.set_zlabel('Summary statistic')
        ax.scatter(self.best_nb_neighbors_ss, self.best_std_dev_ss, self.best_summary_statistic, s=50, c='r', marker = '*', zorder = 100)
        ax.view_init(azim=60, elev=20)  
        ax.xaxis.set_major_locator(plt.MaxNLocator(5))
        ax.zaxis.set_major_locator(plt.MaxNLocator(5))  
            
        #Graphing filtered point cloud using the selected method
            
        ax = fig.add_subplot(2,4,5, projection='3d')
        ax.set_title('Filtered Point Cloud\n with ' + method + ' Method')
        if method == 'PCA':
            cleaned_point_cloud_points = np.asarray(self.best_PCA_cleaned_cloud.points)
            ax.scatter(cleaned_point_cloud_points[:,0], cleaned_point_cloud_points[:,1], cleaned_point_cloud_points[:,2], s=.5)
            ax.set_aspect('equal', adjustable='box')
            ax.set_yticklabels([])
            ax.set_xticklabels([])
            ax.set_zticklabels([])
            ax.view_init(azim=70, elev=20)
            
            
        elif method == 'Average distance':
            cleaned_point_cloud_points = np.asarray(self.best_average_distance_cleaned_cloud.points)
            ax.scatter(cleaned_point_cloud_points[:,0], cleaned_point_cloud_points[:,1], cleaned_point_cloud_points[:,2], s=.5)
            ax.set_aspect('equal', adjustable='box')
            ax.set_yticklabels([])
            ax.set_xticklabels([])
            ax.set_zticklabels([])
            ax.view_init(azim=70, elev=20)
        
        #Graphing filtered point cloud using the optimal parameters for the summary statistic
        ax = fig.add_subplot(2,4,6, projection='3d')
        ax.set_title('Best Filtered Point Cloud\n by the Summary Statistic')
        cleaned_point_cloud_points = np.asarray(self.optimal_filtered_cloud_ss.points)
        ax.scatter(cleaned_point_cloud_points[:,0], cleaned_point_cloud_points[:,1], cleaned_point_cloud_points[:,2], s=.5)
        ax.set_aspect('equal', adjustable='box')
        ax.set_yticklabels([])
        ax.set_xticklabels([])
        ax.set_zticklabels([])
        ax.view_init(azim=70, elev=20)
        
        #Graphing filtered point cloud using the optimal parameters for the chamfer distance
        ax = fig.add_subplot(2,4,7, projection='3d')
        ax.set_title('Best Filtered Point Cloud\n by the Chamfer Distance')
        cleaned_point_cloud_points = np.asarray(self.optimal_filtered_cloud_chamfer_distance.points)
        ax.scatter(cleaned_point_cloud_points[:,0], cleaned_point_cloud_points[:,1], cleaned_point_cloud_points[:,2], s=.5)
        ax.set_aspect('equal', adjustable='box')
        ax.set_yticklabels([])
        ax.set_xticklabels([])
        ax.set_zticklabels([])
        ax.view_init(azim=70, elev=20)
        
        #Graphing chamfer distance over input domain
            
        ax = fig.add_subplot(2,4,8, projection='3d', computed_zorder=False)
        ax.plot_surface(self.nb_neighbor_points_pca_chamfer_distance, self.std_dev_points_pca_chamfer_distance, np.transpose(self.chamfer_distance_data) * 10**4,
                            cmap=cm.coolwarm,
                            linewidth=0, antialiased=False)
        ax.set_title('Chamfer Distance over Input Domain')
        ax.set_xlabel('Number of neighbors')
        ax.set_ylabel('Standard deviation')
        ax.set_zlabel('Chamfer distance * 1e4')
        ax.scatter(self.best_nb_neighbors_chamfer_distance, self.best_std_dev_chamfer_distance, self.best_chamfer_distance * 10**4, s=50, c='r', marker = '*', zorder = 100)
        ax.view_init(azim=140, elev=20)    
        ax.yaxis.labelpad=labelpad
        ax.xaxis.labelpad=labelpad
        ax.zaxis.labelpad=labelpad
        ax.xaxis.set_major_locator(plt.MaxNLocator(5))
        ax.zaxis.set_major_locator(plt.MaxNLocator(5))
        
        
        guessed_summary_statistic = self.evaluate_filter_parameters(evaluation_method="Summary statistic", method=method, reduce_to_summary_statistic=True)
        guessed_chamfer_distance = self.evaluate_filter_parameters(evaluation_method="Chamfer distance", method=method, reduce_to_summary_statistic=False)
        if method == 'PCA':
            self.PCA_recall, self.PCA_accuracy, self.PCA_precision, self.PCA_RMSE = self.evaluate_filter_parameters(evaluation_method="Summary statistic", method=method, reduce_to_summary_statistic=False)
        elif method == 'Average distance':
            self.Average_distance_recall, self.Average_distance_accuracy, self.Average_distance_precision, self.Average_distance_RMSE = self.evaluate_filter_parameters(evaluation_method="Summary statistic", method=method, reduce_to_summary_statistic=False)
        summary_statistic_percent_error = ((self.best_summary_statistic - guessed_summary_statistic) * 100) / (self.best_summary_statistic)
        chamfer_distance_percent_error = -((self.best_chamfer_distance - guessed_chamfer_distance) * 100) / (self.best_chamfer_distance)
        fig.text(.5, 0, 'Best summary statistic: %.2f\n Guessed summary statistic: %.2f\n Percent error: %f%%' % (self.best_summary_statistic, guessed_summary_statistic, summary_statistic_percent_error), ha='center', va='bottom')
        fig.text(.5, 1, 'Best chamfer distance * 10e4: %.2f\n Guessed chamfer distance * 10e4: %.2f\n Percent error: %f%%' % (self.best_chamfer_distance * 10**4, guessed_chamfer_distance * 10**4, chamfer_distance_percent_error), ha='center', va='top')
        if method == 'PCA':
            self.PCA_summary_statistic_error = summary_statistic_percent_error
        elif method == 'Average distance':
            self.average_distance_summary_statistic_error = summary_statistic_percent_error
        if method == 'PCA':
            self.PCA_chamfer_distance_error = chamfer_distance_percent_error
        elif method == 'Average distance':
            self.average_distance_chamfer_distance_error = chamfer_distance_percent_error
        
        plt.subplots_adjust(left=0.06, right=0.94, top=0.9, bottom=.06, wspace = .6, hspace = .46)
        plt.gcf().set_size_inches(14, 8)
        
        if show:
            plt.show()
            
        if save_path:
            if not os.path.exists(os.path.join(save_path, 'summary_subplots', name)):
                os.makedirs(os.path.join(save_path, 'summary_subplots', name))
            if method == 'PCA':
                self.PCA_summary_subplots_filepath = os.path.join(os.getcwd(), save_path, 'summary_subplots', name, 'summary_figure_' + f"{method}_{number}" + '.png')
            elif method == 'Average distance':
                method = 'average_distance'
                self.average_distance_summary_subplots_filepath = os.path.join(os.getcwd(), save_path, 'summary_subplots', name, 'summary_figure_' + f"{method}_{number}" + '.png')
            plt.savefig(os.path.join(save_path, 'summary_subplots', name, 'summary_figure_' + f"{method}_{number}"), dpi=300)

        plt.close()

    def transform_noisy_cloud(self, transformation:list[float] = [[0.0, 0.0, 1.0, 0.0], [1.0, 0.0, 0.0, 0.0],
            [0.0, 1.0, 0.0, 0.0], [0.0, 0.0, 0.0, 1.0]], show: bool = False) -> None:
        """Transforms the noisy cloud using the provided transformation matrix.

        Args:
            transformation (list[float], optional): Transformation matrix to use. Defaults to [[0.0, 0.0, 1.0, 0.0], [1.0, 0.0, 0.0, 0.0],
            [0.0, 1.0, 0.0, 0.0], [0.0, 0.0, 0.0, 1.0]].
            show (bool, optional): Whether or not to show the transformed point cloud. Defaults to False.
            
        Raises:
            Exception: If you try to run this function without generating a noisy cloud first, it will raise an exception.
        """      
        if self.noisy_cloud == None:
            raise Exception("No noisy cloud has been generated. Please generate a noisy cloud before calling\
                        this function, by running PointCloud.generate_gaussian_noise.")
          
        self.noisy_cloud = self.noisy_cloud.transform(transformation)
        if show:
            o3d.visualization.draw_geometries([self.noisy_cloud, self.PointCloud])
        return
    
    def clean_noisy_cloud(self, nb_neighbors: int = 0, std_ratio: float = 0, method:str = None, show:bool = False) -> None:
        """Cleans the noisy cloud using the provided parameters. If you provide a method instead of nb_neighbord or std_ratio, 
        it will automatically select the hypotehtically "best" parameters for that method. These "best" parameters are parameters
        that provide a minimum for the method selected.

        Args:
            nb_neighbors (int, optional): nb_neighbors argument to use for the statistical outlier removal filter. Defaults to 0.\n
            std_ratio (int, optional): std_ratio argument to use for the statistical outlier removal filter. Defaults to 0.\n
            method (str, optional): Use either 'PCA' or 'Average distance'. This will automatically select the hypothetically 'best'
            filter parameters using the selected method. Defaults to None.\n
            show_images (bool, optional): Whether or not to show the noisy and filtered point clouds. Defaults to False.
            
        Raises:
            Exception: If you run this function without providing either nb_neighbors and std_ratio or method, it will raise an exception.
        """ 
               
        if (nb_neighbors == 0 or std_ratio == 0) and method == None:
            raise Exception('Please provide either nb_neighbors and std_ratio or method for filter evaluation.')
        
        if method == 'PCA':
            nb_neighbors = self.best_nb_neighbors_PCA
            std_ratio = self.best_std_dev_PCA
            
        elif method == 'Average distance':
            nb_neighbors = self.best_nb_neighbors_mean
            std_ratio = self.best_std_dev_mean
        
        pcd_noisy = copy.copy(self.noisy_cloud)
        self.cleaned_cloud, ind = pcd_noisy.remove_statistical_outlier(nb_neighbors, std_ratio)
        noise_indeces = []
        removed_points = 0
        for point in range(self.original_number_of_points):
            if point not in ind:
                removed_points += 1
            elif point in self.noise_indeces:
                noise_indeces.append(point - removed_points)
        if show == True:
            print("Showing noisy point cloud")
            self.display_noise_not_noise(self.noisy_cloud, self.noise_indeces)
            print("Showing filtered point cloud")
            self.display_noise_not_noise(self.cleaned_cloud, noise_indeces)
    
    def align_cleaned_cloud_and_original_cloud(self, voxel_size: float = .05, show:bool = False) -> None:
        """Aligns the cleaned cloud and the original cloud using the ICP algorithm. This method is taken directly from Open3D's documentation
        (http://www.open3d.org/docs/release/tutorial/pipelines/global_registration.html). 
        R. Rusu, N. Blodow, and M. Beetz, Fast Point Feature Histograms (FPFH) for 3D registration, ICRA, 2009. for details about fpfh cloud.
        
        Args: 
            voxel_size (float, optional): Voxel size to use for downsampling. Defaults to .05.
            show (bool, optional): Whether or not to show the aligned point clouds. Defaults to False.
        
        Raises:
            Exception: If you try to run this function without cleaning the noisy cloud first, it will raise an exception.
        """
        
        if self.cleaned_cloud == None:
            raise Exception("You haven't cleaned the noisy point cloud yet, so there's no clean point cloud to align.")
        
        def draw_registration_result(source, target, transformation):
            source_temp = copy.deepcopy(source)
            target_temp = copy.deepcopy(target)
            source_temp.paint_uniform_color([1, 0, 0])
            target_temp.paint_uniform_color([0, 1, 0])
            source_temp.transform(transformation)
            o3d.visualization.draw_geometries([source_temp, target_temp])
                
        def preprocess_point_cloud(pcd, voxel_size):
            pcd_down = pcd.voxel_down_sample(voxel_size)
            radius_normal = voxel_size * 2
            pcd_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=100))
            radius_feature = voxel_size * 2
            pcd_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
                pcd_down,
                o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=200))
            return pcd_down, pcd_fpfh
        
        
        def execute_global_registration(cleaned_down, original_down, cleaned_fpfh,
                                original_fpfh, voxel_size):
            distance_threshold = voxel_size
            result = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
                cleaned_down, original_down, cleaned_fpfh, original_fpfh, True,
                distance_threshold,
                o3d.pipelines.registration.TransformationEstimationPointToPoint(False),
                3, [
                    o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(
                        0.9),
                    o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(
                        distance_threshold)
                ], o3d.pipelines.registration.RANSACConvergenceCriteria(100000, 0.9999999))
            return result
            
        def refine_registration(source, target, voxel_size):
            radius_normal = voxel_size * 2
            target.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=25))
            source.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=25))
            distance_threshold = voxel_size * 0.2
            result = o3d.pipelines.registration.registration_icp(
                source, target, distance_threshold, result_ransac.transformation,
                o3d.pipelines.registration.TransformationEstimationPointToPlane())
            return result
        
        original_down, original_fpfh = preprocess_point_cloud(self.PointCloud, voxel_size)
        cleaned_down, cleaned_fpfh = preprocess_point_cloud(self.cleaned_cloud, voxel_size)
        
        result_ransac = execute_global_registration(cleaned_down, original_down,
                                            cleaned_fpfh, original_fpfh,
                                            voxel_size)
        
        result_icp = refine_registration(self.cleaned_cloud, self.PointCloud, voxel_size)

        del self.cleaned_cloud.normals[:]
        del self.PointCloud.normals[:]
        
        if show:
            print ("Showing allignment reslts. Red points are the cleaned cloud, and green points are the original cloud.")
            draw_registration_result(self.cleaned_cloud, self.PointCloud, result_icp.transformation)
            
    def insert_custom_noisy_point_cloud(self, noisy_cloud_filepath: str = None, noisyPointCloud: o3d.geometry.PointCloud = None):
        """Inserts a custom noisy point cloud into the PointCloud object. Either provide a filepath or a point cloud. 

        Args:
            noisy_cloud_filepath (str, optional): Filepath to the noisy point cloud. Defaults to None.
            noisyPointCloud (o3d.geometry.PointCloud, optional): Point cloud to use as the noisy point cloud. Defaults to None.

        Raises:
            Exception: If you try to run this function without providing either a filepath or a point cloud, it will raise an exception. It will also raise an exception
            if you try to provide both a filepath and a point cloud.
        """
        
        if (noisy_cloud_filepath == None and noisyPointCloud == None) or (noisy_cloud_filepath != None and noisyPointCloud != None):
            raise Exception("Please provide either a filepath or a point cloud.")
        
        if noisyPointCloud != None:
            self.noisy_cloud = noisyPointCloud
            
        if noisy_cloud_filepath != None:
            self.noisy_cloud = o3d.io.read_point_cloud(noisy_cloud_filepath)
            
    def record_CSV_data(self, dataframe: pd.DataFrame, method: str, sig_figs: int = 3):
        """Records the data from the point cloud into a pandas dataframe.

        Args:
            dataframe (pd.DataFrame): Pandas dataframe to record the data into.
            method (str): Method used to filter the point cloud. Either 'PCA' or 'Average distance'.
            sig_figs (int, optional): Number of significant figures to round the data to. Defaults to 3.
        """
        if method == 'PCA':        
            new_row = {'Summary figure filepath': self.PCA_summary_subplots_filepath, 
                       'Average distance to nearest neighbor in original cloud': self.original_cloud_average_distance, 
                       'Smallest ratio of PCA variance in original cloud': self.original_lowest_variance_ratio,
                       'Filter RMSE': self.PCA_RMSE, 
                       'Filter Recall': self.PCA_recall, 'Filter Accuracy': self.PCA_accuracy, 'Filter Precision': self.PCA_precision, 
                       'Percent error in summary statistic': self.PCA_summary_statistic_error, 'Percent error in chamfer distance': self.PCA_chamfer_distance_error, 'Points in original cloud': self.original_number_of_points, 
                       'Points in optimal filtered cloud': self.best_PCA_filter_number_of_points, 'Original range x': self.original_ranges[0], 
                       'Filtered range x': self.PCA_filtered_ranges[0], 'Original range y': self.original_ranges[1],
                       'Filtered range y': self.PCA_filtered_ranges[1], 'Original range z': self.original_ranges[2], 'Filtered range z': self.PCA_filtered_ranges[2], 
                        'Filter Method': 'PCA'}
        elif method == 'Average distance':
            new_row = {'Summary figure filepath': self.average_distance_summary_subplots_filepath, 
                       'Average distance to nearest neighbor in original cloud': self.original_cloud_average_distance, 
                       'Smallest ratio of PCA variance in original cloud': self.original_lowest_variance_ratio,
                       'Filter RMSE': self.Average_distance_RMSE, 
                       'Filter Recall': self.Average_distance_recall, 'Filter Accuracy': self.Average_distance_accuracy, 'Filter Precision': self.Average_distance_precision, 
                       'Percent error in summary statistic': self.average_distance_summary_statistic_error,
                       'Percent error in chamfer distance': self.average_distance_chamfer_distance_error, 'Points in original cloud': self.original_number_of_points, 
                       'Points in optimal filtered cloud': self.best_average_distance_filter_number_of_points, 'Original range x': self.original_ranges[0], 
                       'Filtered range x': self.Average_distance_filtered_ranges[0],  'Original range y': self.original_ranges[1],
                       'Filtered range y': self.Average_distance_filtered_ranges[1], 'Original range z': self.original_ranges[2],
                       'Filtered range z': self.Average_distance_filtered_ranges[2], 'Filter Method': 'Average distance'}
        
        for key in new_row:
            if not isinstance(new_row[key], str):
                new_row[key] = round(new_row[key], sig_figs)
        
        if dataframe.empty:
            dataframe = pd.DataFrame([new_row])
        else:
            dataframe = pd.concat([dataframe, pd.DataFrame([new_row])], ignore_index=True)
            
        return dataframe
          
    def generate_bounding_box_information(self):
        """Generates the bounding box information for the point cloud.
        """
        Average_distance_filtered_max_bounds = self.best_average_distance_cleaned_cloud.get_max_bound()
        Average_distance_filtered_min_bounds = self.best_average_distance_cleaned_cloud.get_min_bound()
        PCA_filtered_max_bounds = self.best_PCA_cleaned_cloud.get_max_bound()
        PCA_filtered_min_bounds = self.best_PCA_cleaned_cloud.get_min_bound()
        original_max_bounds = self.PointCloud.get_max_bound()
        original_min_bounds = self.PointCloud.get_min_bound()
        
        self.Average_distance_filtered_ranges = np.asarray(Average_distance_filtered_max_bounds) - np.asarray(Average_distance_filtered_min_bounds)
        self.PCA_filtered_ranges = np.asarray(PCA_filtered_max_bounds) - np.asarray(PCA_filtered_min_bounds)
        self.original_ranges = np.asarray(original_max_bounds) - np.asarray(original_min_bounds)
        
        
        return     
    
class PointCloudDatabase:
    
    def __init__(self, point_cloud_folder: str, noise_mean: float = 0, std_dev_distance_mult: float = 3, noise_percent: float = .2, nb_neighbors_upper: int = 100,\
        nb_neighbors_samples: int = 10, std_dev_upper: float = 3, std_dev_samples: int = 5, nb_neighbors_lower: int = 1, std_dev_lower:float = .00001) -> None:
        
        self.point_cloud_folder = point_cloud_folder
        self.noise_mean = noise_mean
        self.std_dev_distance_mult = std_dev_distance_mult
        self.noise_percent = noise_percent
        self.nb_neighbors_upper = nb_neighbors_upper
        self.nb_neighbors_samples = nb_neighbors_samples
        self.std_dev_upper = std_dev_upper
        self.std_dev_samples = std_dev_samples
        self.nb_neighbor_lower = nb_neighbors_lower
        self.std_dev_lower = std_dev_lower
        
    def record_point_cloud_information(self, folder_path: str, runs_per_cloud: int = 10, clouds_to_select = 30) -> None:
        #random.seed(datetime.now().timestamp())
        
        point_cloud_files = glob.glob(os.path.join(self.point_cloud_folder, '*.ply'))
        
        point_cloud_files = list(point_cloud_files)
        random.shuffle(point_cloud_files)
        
        #interesting_points_clouds = ['pcd98.ply', 'pcd157.ply', 'pcd2078.ply']
        
        #point_cloud_files = [x for x in point_cloud_files if os.path.basename(os.path.normpath(x)) in interesting_points_clouds]
               
        if not os.path.exists(folder_path): 
            os.makedirs(folder_path)
        
        greenFill = styles.PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')
        redFill = styles.PatternFill(start_color='FF0000',
            end_color='FF0000',
            fill_type='solid')
        yellowFill = styles.PatternFill(start_color='FFFF00',
            end_color='FFFF00',
            fill_type='solid')
        
        if not os.path.exists(os.path.join(folder_path, "Point_cloud_data.xlsx")):
            wb = Workbook()
            wb.save(os.path.join(folder_path, "Point_cloud_data.xlsx"))
        
        for point_cloud_file in point_cloud_files[0:clouds_to_select]:
            point_cloud_dataframe = pd.DataFrame(columns = ['Summary figure filepath',
                       'Average distance to nearest neighbor in original cloud', 
                       'Smallest ratio of PCA variance in original cloud',
                        'Filter RMSE', 'Filter Recall', 'Filter Accuracy', 'Filter Precision', 
                        'Percent error in summary statistic', 'Percent error in chamfer distance', 
                        'Points in original cloud', 'Points in optimal filtered cloud', 'Original range x',
                        'Filtered range x', 'Original range y', 'Filtered range y', 'Original range z', 'Filtered range z', 'Filter Method'])
            for i in range(runs_per_cloud):
                pointCloudData = PointCloudStructure(pointCloudFilePath=str(point_cloud_file))
                pointCloudData.generate_gaussian_noise(self.noise_mean, self.std_dev_distance_mult, self.noise_percent)
                pointCloudData.generate_statistical_removal_and_PCA_variance_data(self.nb_neighbors_upper, self.nb_neighbors_samples, self.std_dev_upper, self.std_dev_samples, \
                    self.nb_neighbor_lower, self.std_dev_lower)
                pointCloudData.generate_statistical_removal_and_average_distance_data(self.nb_neighbors_upper, self.nb_neighbors_samples, self.std_dev_upper, self.std_dev_samples, \
                    self.nb_neighbor_lower, self.std_dev_lower)
                pointCloudData.generate_summary_statistic_data(100, 64, 3, 64, 1, .00001)
                pointCloudData.generate_chamfer_distance_data(100, 64, 3, 64, 1, .00001)
                pointCloudData.create_summary_subplots(method='Average distance', save_path = folder_path, name=os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0])), number=i+1)
                pointCloudData.create_summary_subplots(method='PCA', save_path = folder_path, name=os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0])), number=i+1)
                pointCloudData.generate_bounding_box_information()
                point_cloud_dataframe = pointCloudData.record_CSV_data(point_cloud_dataframe, 'PCA')
                point_cloud_dataframe = pointCloudData.record_CSV_data(point_cloud_dataframe, 'Average distance')
            point_cloud_dataframe = point_cloud_dataframe.round(2)
            point_cloud_dataframe.loc['PCA mean'] = point_cloud_dataframe.loc[point_cloud_dataframe['Filter Method'] == 'PCA'].mean(numeric_only=True)
            point_cloud_dataframe.loc['PCA mean', 'Summary figure filepath'] = 'PCA mean'
            point_cloud_dataframe.loc['Average Distance mean'] = point_cloud_dataframe.loc[point_cloud_dataframe['Filter Method'] == 'Average distance'].mean(numeric_only=True)
            point_cloud_dataframe.loc['Average Distance mean', 'Summary figure filepath'] = 'Average Distance mean'
            if point_cloud_dataframe.loc['PCA mean', 'Percent error in summary statistic'] < 30:
                PCA_summary_statistic_color = greenFill
            elif point_cloud_dataframe.loc['PCA mean', 'Percent error in summary statistic'] < 100:
                PCA_summary_statistic_color = yellowFill
            else:
                PCA_summary_statistic_color = redFill
            if point_cloud_dataframe.loc['PCA mean', 'Percent error in chamfer distance'] < 30:
                PCA_chamfer_distance_color = greenFill
            elif point_cloud_dataframe.loc['PCA mean', 'Percent error in chamfer distance'] < 100:
                PCA_chamfer_distance_color = yellowFill
            else:
                PCA_chamfer_distance_color = redFill
            if point_cloud_dataframe.loc['Average Distance mean', 'Percent error in summary statistic'] < 30:
                Average_distance_summary_statistic_color = greenFill
            elif point_cloud_dataframe.loc['Average Distance mean', 'Percent error in summary statistic'] < 100:
                Average_distance_summary_statistic_color = yellowFill
            else:
                Average_distance_summary_statistic_color = redFill
            if point_cloud_dataframe.loc['Average Distance mean', 'Percent error in chamfer distance'] < 30:
                Average_distance_chamfer_distance_color = greenFill
            elif point_cloud_dataframe.loc['Average Distance mean', 'Percent error in chamfer distance'] < 100:
                Average_distance_chamfer_distance_color = yellowFill
            else:
                Average_distance_chamfer_distance_color = redFill
            with pd.ExcelWriter(os.path.join(folder_path, "Point_cloud_data.xlsx"), engine="openpyxl", mode = 'a', if_sheet_exists = 'overlay') as writer:
                point_cloud_dataframe.to_excel(writer, sheet_name=os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0])), index=False)
            workbook = load_workbook(os.path.join(folder_path, "Point_cloud_data.xlsx"))
            workbook[os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0]))][f'H{(runs_per_cloud*2)+2}'].fill = PCA_summary_statistic_color
            workbook[os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0]))][f'I{(runs_per_cloud*2)+2}'].fill = PCA_chamfer_distance_color
            workbook[os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0]))][f'H{(runs_per_cloud*2)+3}'].fill = Average_distance_summary_statistic_color
            workbook[os.path.basename(os.path.normpath(os.path.splitext(point_cloud_file)[0]))][f'I{(runs_per_cloud*2)+3}'].fill = Average_distance_chamfer_distance_color
            wb._named_styles['Normal'].number_format = '0.00'
            workbook.save(os.path.join(folder_path, "Point_cloud_data.xlsx"))
            
    





